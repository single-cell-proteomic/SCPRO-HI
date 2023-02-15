from datetime import datetime
import numpy as np
from openpyxl import load_workbook
import pandas as pd
import plotly.express as px
from plotly.subplots import make_subplots
import scib
import scanpy as sc
from scib.metrics import lisi
from sklearn.cluster import KMeans
from sklearn.metrics import silhouette_score, adjusted_rand_score, normalized_mutual_info_score
import umap.umap_ as umap

def umap_projection(X, Y, name):
    if isinstance(X, pd.DataFrame) == False:
        df = pd.DataFrame(X)
    else:
        df = X.copy()
    category_o = {}
    for lbl in np.unique(Y):
        category_o[lbl] = np.count_nonzero(Y == lbl)
    umap_2d = umap.UMAP(n_components=2, init='random', random_state=0)
    proj_2d = umap_2d.fit_transform(df)
    df[name] = list(Y)
    # print(type(proj_2d))
    fig_2d = px.scatter(
      proj_2d, x=0, y=1, opacity=0.5,
      color=df[name], labels = {'color': name} 
    )
    fig_2d.show()

class EVAL():
    Title = ""
    Owner = "Researcher"
    file_path = "Results.xlsx"
    annData = None
    annData_2 = None
    DB = ""
    DB_2 = ""
    Date = "" 
    cluster = ""
    predicted = ""

    def __init__(self, _title, annData, clust, preds, i_method, annData_2 = None, exp = "All", save = True, verbose = False):
        self.Title = _title
        self.Date = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        self.DB = annData.uns["name"]
        self.annData = annData
        self.cluster = clust
        self.predicted = preds
        self.i_method = i_method
        self.results = {}
        if self.cluster not in self.annData.obs:
            self.annData.obsm["umap_features"] = umap.UMAP(random_state = 42, n_components = 2).fit_transform(self.annData.X)    
            model = DBSCAN(eps = 0.30, min_samples = 3).fit(self.annData.obsm["umap_features"])
            cluster_ids = [str(x) for x in model.labels_]
            self.annData.obs[self.cluster] = cluster_ids
        if exp == "All":
            sc.pp.neighbors(self.annData)
            sc.pp.pca(self.annData)
            self.results["SS_C"] = self._SS_C()
            self.results["SS"] = self._SS()
            if verbose:
                print("SS is calculated = ", self.results["SS"])
            self.results["ARI"] = self._ARI()
            if verbose:
                print("ARI is calculated = ", self.results["ARI"])
            self.results["iLisi"] = self._ilisi()
            if verbose:
                print("iLisi is calculated = ", self.results["iLisi"])
            self.results["cLisi"] = self._clisi()
            if verbose:
                print("clisi is calculated = ", self.results["cLisi"])
            if annData_2 is not None:
                self.annData_2 = annData_2
                self.DB_2 = annData_2.uns["name"]
                self.results["HVG"] = self._HVG()
                if verbose:
                    print("HVG is calculated = ", self.results["HVG"])
        else:
            exec("self." + exp + " = self._" + exp + "()")
            if verbose:
                exec("print ('" +exp + " is calculated = ', self." + exp + ")")
        if save:
            self.SAVE()
    def get_results(self):
        return self.results
    
    def _SS(self):
        return silhouette_score(self.annData.X, self.annData.obs[self.cluster])
    
    def _SS_C(self):
        kmeans = KMeans(n_clusters=7)
        kmeans.fit(self.annData.X)
        return silhouette_score(self.annData.X, kmeans.labels_)
    
    def _ARI(self):
        return adjusted_rand_score(self.annData.obs[self.cluster], self.annData.obs[self.predicted])

    def _ilisi(self):
        self.annData.obs["batch_id"] = self.annData.obs["batch_id"].astype("category")
        return lisi.ilisi_graph(self.annData, "batch_id", type_= "full", n_cores=16, verbose=True)

    def _clisi(self):
        self.annData.obs[self.cluster] = self.annData.obs[self.cluster].astype("category")
        return scib.metrics.clisi_graph(self.annData, "batch_id", self.cluster, k0 = 15)

    def _HVG(self):
        self.annData.obs["batch_id"] = self.annData.obs["batch_id"].astype("category")
        self.annData_2.obs["batch_id"] = self.annData_2.obs["batch_id"].astype("category")
        return scib.metrics.hvg_overlap(self.annData, self.annData_2, "batch_id", n_hvg = self.annData.shape[1])

    def SAVE(self):
        wb = load_workbook(filename = self.file_path)
        ws = wb['Sayfa1']
        row = [self.Title,
        self.Owner,
        self.DB,
        self.DB_2,
        self.Date]
        metric_list = ["SS_C", "SS", "ARI", "iLisi", "cLisi", "HVG"]
        for metric in metric_list:
            if metric in self.results:
                row.append(self.results[metric])
            else:
                row.append("-")
        newRowLocation = len(ws['A']) + 1
        for col in range(1, len(row) + 1):
            ws.cell(column = col, row = newRowLocation, value = row[col-1])
        wb.save(filename = self.file_path)
        wb.close()